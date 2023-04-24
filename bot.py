# Imports
from email.mime.application import MIMEApplication
from botcity.web import WebBot, Browser, By
import openpyxl
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage

# Import for integration with BotCity Maestro SDK
from botcity.maestro import *

# Disable errors if we are not connected to Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False


def main():
    # Runner passes the server url, the id of the task being executed,
    # the access token and the parameters that this task receives (when applicable).
    maestro = BotMaestroSDK.from_sys_args()
    ## Fetch the BotExecution with details from the task, including parameters
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    bot = WebBot()

    # Configure whether or not to run on headless mode
    bot.headless = False

    # Uncomment to change the default Browser to Firefox
    bot.browser = Browser.CHROME

    # Uncomment to set the WebDriver path
    bot.driver_path = "./chromedriver.exe"

    # Creates the excel file
    # Cria tabela excel
    workbook = openpyxl.Workbook()    
    ws = workbook.active
    ws["A1"] = "Estado"
    ws["B1"] = "Gentílico"
    ws["C1"] = "Capital"
    ws["D1"] = "Governador"
    ws["E1"] = "População"
    ws["F1"] = "IDH"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8    
    
    # Web Scrap (function to gather the IBGE data)
    # Web Scrap (função para colher os dados específicos)
    def buscar_dados(site_desejado):
      
        bot.browse(site_desejado)
        # States
        # Estado
        if not bot.find_text( "Estado1", threshold=230, waiting_time=10000):
            not_found("Estado1")
        bot.triple_click_relative(21, 36)
        bot.control_c()
        bot.wait(500)
        estado1 = str(bot.get_clipboard())
                
        # Gentile
        # Gentilico 
        if not bot.find_text( "Gentilico1", threshold=230, waiting_time=10000):
            not_found("Gentilico1")
        bot.triple_click_relative(25, 26)
        bot.control_c()
        bot.wait(500)
        gentilico1 = str(bot.get_clipboard())
                        
        # Capital
        # Capital
        if not bot.find_text( "Capital1", threshold=230, waiting_time=10000):
            not_found("Capital1")
        bot.triple_click_relative(36, 25)
        bot.wait(500)
        bot.control_c()
        capital1 = str(bot.get_clipboard())
                    
        # Governor
        # Governador
        if not bot.find_text( "governador1", threshold=230, waiting_time=10000):
            not_found("governador1")
        bot.triple_click_relative(90, 23)
        bot.wait(500)
        bot.control_c()
        governador1 = str(bot.get_clipboard())
                            
        # Estimate population
        # População estimada
        if not bot.find_text( "populacao_estimada1", threshold=230, waiting_time=10000):
            not_found("populacao_estimada1")
        bot.double_click_relative(335, 9)
        bot.wait(500)
        bot.control_c()
        população_estimada1 = str(bot.get_clipboard())
                        
        # IDH    
        bot.scroll_down(clicks=27)
        if not bot.find_text( "Economia1", threshold=230, waiting_time=10000):
            not_found("Economia1")
        bot.double_click_relative(55, 144)
        bot.wait(500)
        bot.control_c()
        IDH1 = str(bot.get_clipboard())
        bot.stop_browser()

        # Excel
        ws.append([estado1, gentilico1, capital1,governador1,população_estimada1,IDH1])

    # List with the websites URL(States)
    # Lista de URL dos sites(estados)
    lista_sites = ["https://cidades.ibge.gov.br/brasil/ac/panorama", "https://cidades.ibge.gov.br/brasil/al/panorama", "https://cidades.ibge.gov.br/brasil/ap/panorama", "https://cidades.ibge.gov.br/brasil/am/panorama", "https://cidades.ibge.gov.br/brasil/ba/panorama", "https://cidades.ibge.gov.br/brasil/ce/panorama", "https://cidades.ibge.gov.br/brasil/df/panorama", "https://cidades.ibge.gov.br/brasil/es/panorama", "https://cidades.ibge.gov.br/brasil/go/panorama", "https://cidades.ibge.gov.br/brasil/ma/panorama", "https://cidades.ibge.gov.br/brasil/mt/panorama", "https://cidades.ibge.gov.br/brasil/ms/panorama", "https://cidades.ibge.gov.br/brasil/mg/panorama", "https://cidades.ibge.gov.br/brasil/pr/panorama", "https://cidades.ibge.gov.br/brasil/pb/panorama", "https://cidades.ibge.gov.br/brasil/pa/panorama", "https://cidades.ibge.gov.br/brasil/pe/panorama", "https://cidades.ibge.gov.br/brasil/pi/panorama", "https://cidades.ibge.gov.br/brasil/rn/panorama", "https://cidades.ibge.gov.br/brasil/rs/panorama", "https://cidades.ibge.gov.br/brasil/rj/panorama", "https://cidades.ibge.gov.br/brasil/ro/panorama", "https://cidades.ibge.gov.br/brasil/rr/panorama", "https://cidades.ibge.gov.br/brasil/sc/panorama", "https://cidades.ibge.gov.br/brasil/se/panorama", "https://cidades.ibge.gov.br/brasil/sp/panorama", "https://cidades.ibge.gov.br/brasil/to/panorama"]

    # 'For' responsable to execute the function to gather the data.
    # 'For' responsável para a execução da função de busca de dados.
    for i in range(len(lista_sites)):
        buscar_dados(lista_sites[i])
        workbook.save("relatorio_IBGE.xlsx")
    
    workbook.save("relatorio_IBGE.xlsx")

    # Sending the e-mail.
    # Enviando e-mail
    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    # Its necessary to provide your informations on the designated fields, then the bot will be able to send the message correctly.
    # Aqui é necessario informar o seu login(email) e senha para o bot enviar a mensagem corretamente.
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    # Remember:  The password does not refer to the email, it is necessary to create a password through the "Apps Passwords" feature, if using google gmail.
    # Lembre-se: A senha não é referente ao e-mail, é necessário criar uma senha através do recurso "Senhas de Apps" no gmail.
    server.login("SEU-EMAIL", "SENHA-APP")
    
    msg = EmailMessage()
    msg.set_content("Segue em anexo dados colhidos através do IBGE referentes a todos os Estados Brasileiros.")
    msg["Subject"] = "Dados IBGE"
    msg["From"] = "SEU-EMAIL"
    msg["To"] = "EMAIL-DO-DESTINATÁRIO"

    with open("relatorio_IBGE.xlsx", "rb") as f:
        arquivo = f.read()
    msg.add_attachment(arquivo, maintype="application", subtype="octet-stream", filename="relatorio_IBGE.xlsx")

    server.send_message(msg)
    server.quit()

    # Uncomment to mark this task as finished on BotMaestro
    # maestro.finish_task(
    #     task_id=execution.task_id,
    #     status=AutomationTaskFinishStatus.SUCCESS,
    #     message="Task Finished OK."
    # )


def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    main()